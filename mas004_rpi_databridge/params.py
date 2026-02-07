import io
import re
from typing import Optional, Dict, Any, Tuple, List

import openpyxl

from mas004_rpi_databridge.db import DB, now_ts

SHEET_NAME = "Parameter"


def _to_str(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


class ParamStore:
    def __init__(self, db: DB):
        self.db = db

    def import_xlsx(self, file_path: str) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f"Excel-Sheet '{SHEET_NAME}' nicht gefunden. Vorhanden: {wb.sheetnames}")
        ws = wb[SHEET_NAME]

        inserted = 0
        updated = 0
        skipped = 0

        headers_raw = [(_to_str(ws.cell(1, c).value) or "").strip() for c in range(1, ws.max_column + 1)]
        header_map: Dict[str, int] = {}
        for i, h in enumerate(headers_raw, start=1):
            nh = _norm_header(h)
            if nh:
                header_map[nh] = i

        def col_any(*names: str) -> Optional[int]:
            for n in names:
                nn = _norm_header(n)
                if nn in header_map:
                    return header_map[nn]
            for want in names:
                want2 = (want or "").strip().lower()
                for idx, h in enumerate(headers_raw, start=1):
                    if (h or "").strip().lower() == want2:
                        return idx
            return None

        c_type = col_any("Params_Type.", "Params_Type.:", "Params Type", "Params_Type")
        c_id   = col_any("Param ID.", "Param. ID.:", "Param. ID.", "Param ID", "Param_ID", "Param. ID")

        c_min   = col_any("Min.", "Min.:", "Min")
        c_max   = col_any("Max.", "Max.:", "Max")
        c_def   = col_any("Default Value", "Default Value:", "Default")
        c_unit  = col_any("Einheit", "Unit")
        c_rw    = col_any("R/W:", "R/W", "RW", "R_W")
        c_dtype = col_any("Data Type", "DataType", "Datatype")
        c_name  = col_any("Name")
        c_fmt   = col_any("Format relevant?", "Format relevant", "Format")
        c_msg   = col_any("Message")
        c_cause = col_any("Possible Cause", "Possible cause")
        c_eff   = col_any("Effects", "Effect")
        c_rem   = col_any("Remedy")

        if not c_type or not c_id:
            raise RuntimeError(
                "Pflichtspalten fehlen. Erwartet (varianten): 'Params_Type' und 'Param ID'. "
                f"Gefunden (normalisiert): {sorted(list(header_map.keys()))}"
            )

        with self.db._conn() as c:
            for r in range(2, ws.max_row + 1):
                ptype = (_to_str(ws.cell(r, c_type).value) or "").strip()
                pid = (_to_str(ws.cell(r, c_id).value) or "").strip()

                if not ptype and not pid:
                    continue
                if not ptype or not pid:
                    skipped += 1
                    continue

                pkey = f"{ptype}{pid}"

                min_v = _to_float(ws.cell(r, c_min).value) if c_min else None
                max_v = _to_float(ws.cell(r, c_max).value) if c_max else None
                default_v = _to_str(ws.cell(r, c_def).value) if c_def else None
                unit = _to_str(ws.cell(r, c_unit).value) if c_unit else None
                rw = _to_str(ws.cell(r, c_rw).value) if c_rw else None
                dtype = _to_str(ws.cell(r, c_dtype).value) if c_dtype else None
                name = _to_str(ws.cell(r, c_name).value) if c_name else None
                fmt = _to_str(ws.cell(r, c_fmt).value) if c_fmt else None
                msg = _to_str(ws.cell(r, c_msg).value) if c_msg else None
                cause = _to_str(ws.cell(r, c_cause).value) if c_cause else None
                eff = _to_str(ws.cell(r, c_eff).value) if c_eff else None
                rem = _to_str(ws.cell(r, c_rem).value) if c_rem else None

                ts = now_ts()

                exists = c.execute("SELECT 1 FROM params WHERE pkey=?", (pkey,)).fetchone() is not None
                if exists:
                    c.execute(
                        """UPDATE params SET
                           ptype=?, pid=?, min_v=?, max_v=?, default_v=?, unit=?, rw=?, dtype=?,
                           name=?, format_relevant=?, message=?, possible_cause=?, effects=?, remedy=?,
                           updated_ts=?
                           WHERE pkey=?""",
                        (ptype, pid, min_v, max_v, default_v, unit, rw, dtype, name, fmt, msg, cause, eff, rem, ts, pkey)
                    )
                    updated += 1
                else:
                    c.execute(
                        """INSERT INTO params(
                           pkey,ptype,pid,min_v,max_v,default_v,unit,rw,dtype,name,format_relevant,
                           message,possible_cause,effects,remedy,updated_ts
                           )
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (pkey, ptype, pid, min_v, max_v, default_v, unit, rw, dtype, name, fmt, msg, cause, eff, rem, ts)
                    )
                    inserted += 1

        return {"ok": True, "inserted": inserted, "updated": updated, "skipped": skipped}

    def get_meta(self, pkey: str) -> Optional[Dict[str, Any]]:
        with self.db._conn() as c:
            row = c.execute(
                """SELECT pkey,ptype,pid,min_v,max_v,default_v,unit,rw,dtype,name,format_relevant,message
                   FROM params WHERE pkey=?""",
                (pkey,)
            ).fetchone()
        if not row:
            return None
        keys = ["pkey","ptype","pid","min_v","max_v","default_v","unit","rw","dtype","name","format_relevant","message"]
        return dict(zip(keys, row))

    def get_value(self, pkey: str) -> Optional[str]:
        with self.db._conn() as c:
            row = c.execute("SELECT value FROM param_values WHERE pkey=?", (pkey,)).fetchone()
        return row[0] if row else None

    def get_effective_value(self, pkey: str) -> str:
        v = self.get_value(pkey)
        if v is not None:
            return v
        meta = self.get_meta(pkey)
        dv = (meta or {}).get("default_v")
        return dv if dv is not None else "0"

    def set_value(self, pkey: str, value: str) -> Tuple[bool, str]:
        meta = self.get_meta(pkey)
        if not meta:
            return False, "NAK_UnknownParam"

        rw = (meta.get("rw") or "").strip().upper()
        if rw == "R":
            return False, "NAK_ReadOnly"

        min_v = meta.get("min_v")
        max_v = meta.get("max_v")
        try:
            fv = float(value)
            if min_v is not None and fv < float(min_v):
                return False, "NAK_OutOfRange"
            if max_v is not None and fv > float(max_v):
                return False, "NAK_OutOfRange"
        except Exception:
            pass

        ts = now_ts()
        with self.db._conn() as c:
            c.execute(
                "INSERT INTO param_values(pkey,value,updated_ts) VALUES(?,?,?) "
                "ON CONFLICT(pkey) DO UPDATE SET value=excluded.value, updated_ts=excluded.updated_ts",
                (pkey, str(value), ts)
            )
            # Fuer den Testbetrieb soll ein erfolgreiches Write auch den Default-Wert fortschreiben.
            c.execute(
                "UPDATE params SET default_v=?, updated_ts=? WHERE pkey=?",
                (str(value), ts, pkey),
            )
        return True, "OK"

    def apply_device_value(self, pkey: str, value: str) -> Tuple[bool, str]:
        """
        Speichert einen von Device-Seite gemeldeten Wert (TTO/Laser/ESP) lokal.
        Im Unterschied zu set_value() wird rw NICHT blockiert, da ReadOnly hier nur
        fuer Mikrotom-Schreibrechte gilt, nicht fuer eingehende Status-/Fehlermeldungen.
        """
        meta = self.get_meta(pkey)
        if not meta:
            return False, "NAK_UnknownParam"

        ts = now_ts()
        with self.db._conn() as c:
            c.execute(
                "INSERT INTO param_values(pkey,value,updated_ts) VALUES(?,?,?) "
                "ON CONFLICT(pkey) DO UPDATE SET value=excluded.value, updated_ts=excluded.updated_ts",
                (pkey, str(value), ts)
            )
        return True, "OK"

    def update_meta(
        self,
        pkey: str,
        default_v: Optional[str] = None,
        min_v: Optional[float] = None,
        max_v: Optional[float] = None,
        rw: Optional[str] = None,
    ) -> Tuple[bool, str]:
        meta = self.get_meta(pkey)
        if not meta:
            return False, "NAK_UnknownParam"

        new_min = meta.get("min_v") if min_v is None else min_v
        new_max = meta.get("max_v") if max_v is None else max_v
        new_def = meta.get("default_v") if default_v is None else str(default_v)
        new_rw = meta.get("rw") if rw is None else str(rw).strip()

        # rw normalisieren
        if new_rw is not None:
            rwu = new_rw.strip().upper()
            if rwu in ("RW", "R_W", "R/W"):
                rwu = "R/W"
            if rwu not in ("R", "W", "R/W", ""):
                return False, "NAK_BadRW"
            new_rw = rwu if rwu else None

        # min/max check
        if new_min is not None and new_max is not None:
            if float(new_min) > float(new_max):
                return False, "NAK_MinGreaterThanMax"

        # default range check (nur wenn numeric)
        try:
            fv = float(new_def) if new_def is not None else None
            if fv is not None:
                if new_min is not None and fv < float(new_min):
                    return False, "NAK_DefaultOutOfRange"
                if new_max is not None and fv > float(new_max):
                    return False, "NAK_DefaultOutOfRange"
        except Exception:
            pass

        with self.db._conn() as c:
            c.execute(
                """UPDATE params
                   SET default_v=?, min_v=?, max_v=?, rw=?, updated_ts=?
                   WHERE pkey=?""",
                (new_def, new_min, new_max, new_rw, now_ts(), pkey)
            )
        return True, "OK"

    def list_params(self, ptype: Optional[str] = None, q: Optional[str] = None, limit: int = 200, offset: int = 0):
        limit = max(1, min(int(limit), 1000))
        offset = max(0, int(offset))
        where = []
        args = []

        if ptype:
            where.append("ptype=?")
            args.append(ptype)

        if q:
            q2 = f"%{q}%"
            where.append("(pkey LIKE ? OR name LIKE ? OR message LIKE ?)")
            args.extend([q2, q2, q2])

        wsql = ("WHERE " + " AND ".join(where)) if where else ""
        sql = f"""SELECT pkey,ptype,pid,min_v,max_v,default_v,unit,rw,dtype,name,message
                  FROM params
                  {wsql}
                  ORDER BY ptype ASC, pid ASC
                  LIMIT ? OFFSET ?"""
        args.extend([limit, offset])

        with self.db._conn() as c:
            rows = c.execute(sql, args).fetchall()

        out = []
        for r in rows:
            pkey = r[0]
            cur = self.get_value(pkey)
            eff = cur if cur is not None else (r[5] if r[5] is not None else "0")
            out.append({
                "pkey": r[0],
                "ptype": r[1],
                "pid": r[2],
                "min_v": r[3],
                "max_v": r[4],
                "default_v": r[5],
                "current_v": cur,
                "effective_v": eff,
                "unit": r[6],
                "rw": r[7],
                "dtype": r[8],
                "name": r[9],
                "message": r[10],
            })
        return out

    def export_xlsx_bytes(self, ptype: Optional[str] = None, q: Optional[str] = None) -> bytes:
        rows = self.list_params(ptype=ptype, q=q, limit=100000, offset=0)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        headers = [
            "Params_Type.:",
            "Param. ID.:",
            "Min.:",
            "Max.:",
            "Default Value",
            "Current Value",
            "Effective Value",
            "Einheit",
            "R/W:",
            "Data Type",
            "Name",
            "Message",
        ]
        ws.append(headers)

        for r in rows:
            ws.append([
                r.get("ptype"),
                r.get("pid"),
                r.get("min_v"),
                r.get("max_v"),
                r.get("default_v"),
                r.get("current_v"),
                r.get("effective_v"),
                r.get("unit"),
                r.get("rw"),
                r.get("dtype"),
                r.get("name"),
                r.get("message"),
            ])

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()
