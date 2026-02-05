import os
import sqlite3
from typing import Optional, Dict, Any, List, Tuple
from contextlib import contextmanager
import openpyxl

def _colnames(cur, table: str) -> List[str]:
    rows = cur.execute(f"PRAGMA table_info({table})").fetchall()
    return [r[1] for r in rows]

class ParamStore:
    def __init__(self, db):
        self.db = db
        self.ensure_schema()

    def ensure_schema(self):
        with self.db._conn() as c:
            c.execute("""
            CREATE TABLE IF NOT EXISTS params(
              ptype TEXT NOT NULL,
              pid   TEXT NOT NULL,
              min_v REAL,
              max_v REAL,
              rw    TEXT,
              default_value TEXT,
              current_value TEXT,
              description TEXT,
              PRIMARY KEY (ptype, pid)
            )""")

    def get(self, ptype: str, pid: str) -> Optional[Dict[str, Any]]:
        with self.db._conn() as c:
            row = c.execute("""
                SELECT ptype,pid,min_v,max_v,rw,default_value,current_value,description
                FROM params WHERE ptype=? AND pid=?""", (ptype, pid)).fetchone()
        if not row:
            return None
        keys = ["ptype","pid","min_v","max_v","rw","default_value","current_value","description"]
        return dict(zip(keys, row))

    def set_current(self, ptype: str, pid: str, value: str):
        with self.db._conn() as c:
            c.execute("""
              UPDATE params SET current_value=? WHERE ptype=? AND pid=?""",
              (str(value), ptype, pid)
            )

    def upsert(self, rec: Dict[str, Any]):
        with self.db._conn() as c:
            c.execute("""
              INSERT INTO params(ptype,pid,min_v,max_v,rw,default_value,current_value,description)
              VALUES(?,?,?,?,?,?,?,?)
              ON CONFLICT(ptype,pid) DO UPDATE SET
                min_v=excluded.min_v,
                max_v=excluded.max_v,
                rw=excluded.rw,
                default_value=excluded.default_value,
                current_value=excluded.current_value,
                description=excluded.description
            """, (
                rec["ptype"], rec["pid"],
                rec.get("min_v"), rec.get("max_v"),
                rec.get("rw"),
                rec.get("default_value"),
                rec.get("current_value"),
                rec.get("description"),
            ))

    def list(self, ptype: Optional[str]=None, limit: int=500) -> List[Dict[str, Any]]:
        with self.db._conn() as c:
            if ptype:
                rows = c.execute("""
                    SELECT ptype,pid,min_v,max_v,rw,default_value,current_value,description
                    FROM params WHERE ptype=? ORDER BY pid LIMIT ?""", (ptype, limit)).fetchall()
            else:
                rows = c.execute("""
                    SELECT ptype,pid,min_v,max_v,rw,default_value,current_value,description
                    FROM params ORDER BY ptype,pid LIMIT ?""", (limit,)).fetchall()
        keys = ["ptype","pid","min_v","max_v","rw","default_value","current_value","description"]
        return [dict(zip(keys, r)) for r in rows]

    # -------- Excel Import --------
    def import_xlsx(self, xlsx_path: str) -> Tuple[int, List[str]]:
        """
        Erwartete Spalten (entweder per Header oder per Position):
        A: PARAMS_Type
        B: PARAM.ID
        C: Min
        D: Max
        E: Default Value
        F: Description (optional)
        G: R/W
        """
        if not os.path.exists(xlsx_path):
            return (0, [f"file not found: {xlsx_path}"])

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # Header erkennen
        first = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        upper = [h.upper() for h in first]
        has_header = any("TYPE" in h or "PARAM" in h or "R/W" in h for h in upper)

        def pick(row, idx):
            return row[idx] if idx < len(row) else None

        start_row = 2 if has_header else 1

        count = 0
        warnings = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            ptype = pick(row, 0)
            pid = pick(row, 1)
            if ptype is None or pid is None:
                continue
            ptype = str(ptype).strip().upper()
            pid = str(pid).strip()
            # keep leading zeros if excel text; if numeric -> no zeros
            # (Normalization übernimmt später protocol.normalize_pid beim Matching)
            min_v = pick(row, 2)
            max_v = pick(row, 3)
            default_v = pick(row, 4)
            desc = pick(row, 5)
            rw = pick(row, 6)

            rec = {
                "ptype": ptype,
                "pid": pid,
                "min_v": float(min_v) if min_v not in (None, "") else None,
                "max_v": float(max_v) if max_v not in (None, "") else None,
                "rw": str(rw).strip().upper() if rw not in (None, "") else None,
                "default_value": str(default_v).strip() if default_v not in (None, "") else None,
                "current_value": str(default_v).strip() if default_v not in (None, "") else None,
                "description": str(desc).strip() if desc not in (None, "") else None,
            }
            self.upsert(rec)
            count += 1

        return (count, warnings)
